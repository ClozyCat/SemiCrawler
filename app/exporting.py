from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import EXPORT_COLUMNS


AUDIT_COLUMNS = [("status", "处理状态"), ("evidence_json", "字段证据"), ("confidence_json", "字段置信度"),
                 ("article_id", "原文ID"), ("task_id", "任务ID"), ("created_at", "开始时间"), ("updated_at", "更新时间")]


def _columns(audit=False):
    return EXPORT_COLUMNS + (AUDIT_COLUMNS if audit else [])


def _rows(records, audit=False):
    for record in records:
        yield [
            value.isoformat() if hasattr(value, "isoformat") else (value or "")
            for field, _ in _columns(audit)
            for value in [getattr(record, field)]
        ]


def make_csv(records, audit=False) -> bytes:
    stream = StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow([label for _, label in _columns(audit)])
    writer.writerows(_rows(records, audit))
    return ("\ufeff" + stream.getvalue()).encode("utf-8")


def make_xlsx(records, audit=False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "结构化结果"
    sheet.append([label for _, label in _columns(audit)])
    for row in _rows(records, audit):
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="DDEBFA")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="17324D")
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [18, 24, 26, 14, 18, 18, 30, 24, 42, 50] + ([16, 50, 50, 12, 12, 22, 22] if audit else [])
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
