from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.constants import EXPORT_COLUMNS
from app.database import SessionLocal
from app.main import _scheduled_task_payload
from app.models import RawArticle, ScheduledTask, SourceVersion, StructuredRecord


def test_defaults_sources_and_meta(client):
    meta = client.get("/api/meta")
    assert meta.status_code == 200
    assert meta.json()["default_start_date"] == "2026-08-01"
    assert "资讯类型" not in meta.json()["info_types"]
    assert "项目立项" in meta.json()["info_types"]
    assert meta.json()["info_types"][:2] == ["项目规划", "项目立项"]

    sources = client.get("/api/sources").json()
    assert [item["name"] for item in sources] == [
        "全球半导体观察（DRAMx）",
        "半导体产业网",
    ]
    assert all(item["enabled"] for item in sources)


def test_source_persistence_and_toggle(client):
    response = client.post(
        "/api/sources",
        json={
            "name": "测试来源",
            "base_url": "https://example.com",
            "enabled": True,
            "config": {
                "entry_urls": ["https://example.com/news"],
                "article_url_pattern": "/news/\\d+",
                "selectors": {
                    "list_links": "a",
                    "title": "h1",
                    "date": ".date",
                    "content": ".content",
                },
            },
        },
    )
    assert response.status_code == 201
    source_id = response.json()["id"]

    updated = client.patch(f"/api/sources/{source_id}", json={"enabled": False})
    assert updated.status_code == 200
    assert updated.json()["enabled"] is False

    deleted = client.delete(f"/api/sources/{source_id}")
    assert deleted.status_code == 200
    assert deleted.json() == {"deleted": 1}
    assert all(item["id"] != source_id for item in client.get("/api/sources").json())
    with SessionLocal() as db:
        assert db.query(SourceVersion).filter_by(source_id=source_id).count() == 0


def test_source_delete_rejects_builtin_and_source_with_articles(client):
    builtin = client.get("/api/sources").json()[0]
    response = client.delete(f"/api/sources/{builtin['id']}")
    assert response.status_code == 409
    assert response.json()["detail"] == "内置信息源不可删除，可将其停用"

    created = client.post(
        "/api/sources",
        json={
            "name": "有数据的来源",
            "base_url": "https://example.org",
            "config": {
                "entry_urls": ["https://example.org/news"],
                "article_url_pattern": "/news/",
                "selectors": {
                    "list_links": "a",
                    "title": "h1",
                    "date": ".date",
                    "content": ".content",
                },
            },
        },
    ).json()
    with SessionLocal() as db:
        db.add(
            RawArticle(
                source_id=created["id"],
                canonical_url="https://example.org/news/1",
                title="测试原文",
                body="正文",
                content_hash="hash",
            )
        )
        db.commit()

    response = client.delete(f"/api/sources/{created['id']}")
    assert response.status_code == 409
    assert response.json()["detail"] == "该信息源已有原始数据，请先删除相关原始数据"


def test_scheduled_task_protects_its_source_and_can_be_managed(client):
    created = client.post(
        "/api/sources",
        json={
            "name": "定时任务测试来源",
            "base_url": "https://example.org",
            "config": {
                "entry_urls": ["https://example.org/news"],
                "article_url_pattern": "/news/",
                "selectors": {"list_links": "a", "title": "h1", "date": ".date", "content": "article"},
                "request": {"rate_limit_per_minute": 10, "timeout_seconds": 10},
            },
        },
    ).json()
    scheduled = client.post(
        "/api/schedules",
        json={
            "name": "每日测试采集",
            "frequency": "daily",
            "hour": 9,
            "start_date": "2026-08-01",
            "source_ids": [created["id"]],
            "keyword_filter_enabled": True,
            "auto_structure_enabled": True,
        },
    )
    assert scheduled.status_code == 201
    saved_schedule = client.get("/api/schedules").json()[0]
    assert saved_schedule["frequency"] == "daily"
    assert saved_schedule["keyword_filter_enabled"] is True
    assert saved_schedule["auto_structure_enabled"] is True

    updated = client.patch(
        f"/api/schedules/{scheduled.json()['id']}",
        json={**scheduled.json(), "keyword_filter_enabled": False},
    )
    assert updated.status_code == 200
    assert updated.json()["keyword_filter_enabled"] is False
    assert updated.json()["auto_structure_enabled"] is True
    with SessionLocal() as db:
        run_payload = _scheduled_task_payload(db.get(ScheduledTask, scheduled.json()["id"]))
    assert run_payload.keyword_filter_enabled is False
    assert run_payload.auto_structure_enabled is True
    assert run_payload.keyword_config is None
    blocked = client.delete(f"/api/sources/{created['id']}")
    assert blocked.status_code == 409
    assert "定时任务" in blocked.json()["detail"]
    assert client.delete(f"/api/schedules/{scheduled.json()['id']}").status_code == 200
    assert client.delete(f"/api/sources/{created['id']}").status_code == 200


def test_web_search_source_uses_simple_natural_language_config(client):
    response = client.post(
        "/api/sources",
        json={
            "name": "联网搜索：先进封装项目",
            "base_url": "https://dokobot.ai",
            "enabled": True,
            "config": {
                "type": "web_search",
                "query": "检索先进封装项目的签约、开工与扩产动态",
                "source_hint": "https://example.gov.cn/news\nhttps://example.com/press",
                "max_results": 20,
            },
        },
    )
    assert response.status_code == 201
    assert response.json()["source_type"] == "web_search"
    assert (
        response.json()["config"]["query"] == "检索先进封装项目的签约、开工与扩产动态"
    )
    assert response.json()["config"]["max_results"] == 20
    assert response.json()["config"]["provider"] == "anysearch"


def test_web_search_source_rejects_invalid_source_hint_lines(client):
    for index, source_hint in enumerate(
        [
            "优先政府园区官网和企业新闻中心",
            "https://one.example/news https://two.example/projects",
        ]
    ):
        response = client.post(
            "/api/sources",
            json={
                "name": f"无效网址来源 {index}",
                "base_url": "https://api.anysearch.com",
                "config": {
                    "type": "web_search",
                    "query": "检索中国半导体项目动态",
                    "source_hint": source_hint,
                },
            },
        )

        assert response.status_code == 422
        assert "每行填写一个有效的 http(s) 网址" in response.json()["detail"]


def test_web_search_source_defaults_to_anysearch(client):
    response = client.post(
        "/api/sources",
        json={
            "name": "Anysearch联网搜索",
            "base_url": "https://api.anysearch.com",
            "config": {
                "type": "web_search",
                "query": "检索中国半导体项目动态",
            },
        },
    )

    assert response.status_code == 201
    assert response.json()["config"]["provider"] == "anysearch"


def test_task_snapshot_and_logs(client, monkeypatch):
    def fake_run(db, task):
        from app.models import TaskLog, utc_now

        task.status = "completed"
        task.completed_at = utc_now()
        db.add(TaskLog(task_id=task.id, message="任务完成：Mock"))
        db.commit()

    monkeypatch.setattr("app.main.run_task", fake_run)
    source_ids = [item["id"] for item in client.get("/api/sources").json()]
    response = client.post(
        "/api/tasks", json={"source_ids": source_ids, "start_date": "2026-08-01"}
    )
    assert response.status_code == 201
    task = response.json()
    assert task["status"] == "queued"
    task = client.get(f"/api/tasks/{task['id']}").json()
    assert task["status"] == "completed"
    assert task["progress"] == 100
    assert task["created_at"].endswith("+08:00")
    assert task["completed_at"].endswith("+08:00")
    assert len(task["source_snapshot"]) == 2

    assert task["keyword_filter_enabled"] is False
    assert task["auto_structure_enabled"] is False

    logs = client.get(f"/api/tasks/{task['id']}/logs").json()
    assert len(logs) == 4
    assert "Mock" in logs[-1]["message"]
    assert all(log["created_at"].endswith("+08:00") for log in logs)


def test_task_snapshots_keyword_and_structure_options(client, monkeypatch):
    monkeypatch.setattr("app.main.run_task", lambda db, task: None)
    source_id = client.get("/api/sources").json()[0]["id"]
    config = [{"industry": "先进封装", "field": "Chiplet", "keywords": "CoWoS、TSV"}]
    response = client.post(
        "/api/tasks",
        json={
            "source_ids": [source_id],
            "start_date": "2026-08-01",
            "keyword_filter_enabled": True,
            "auto_structure_enabled": True,
            "keyword_config": config,
        },
    )
    assert response.status_code == 201
    assert response.json()["keyword_filter_enabled"] is True
    assert response.json()["auto_structure_enabled"] is True
    assert response.json()["keyword_config"] == config


def test_task_uses_saved_keywords_when_legacy_client_omits_config(
    client, monkeypatch
):
    monkeypatch.setattr("app.main.run_task", lambda db, task: None)
    saved_config = {
        "technical": [
            {"industry": "半导体", "field": "集成电路", "keywords": "芯片"}
        ],
        "industry_noun": [
            {"industry": "", "field": "", "keywords": "产业园、二期"}
        ],
        "industry_verb": [
            {"industry": "", "field": "", "keywords": "建设、规划"}
        ],
    }
    setting = client.put(
        "/api/settings/model",
        json={
            "base_url": "https://api.example.com",
            "model_name": "test-model",
            "keyword_config": saved_config,
        },
    )
    assert setting.status_code == 200
    source_id = client.get("/api/sources").json()[0]["id"]

    response = client.post(
        "/api/tasks",
        json={
            "source_ids": [source_id],
            "start_date": "2026-08-01",
            "keyword_filter_enabled": True,
        },
    )

    assert response.status_code == 201, response.text
    assert response.json()["keyword_config"] == saved_config


def test_task_rejects_enabled_filter_with_incomplete_keyword_dimensions(
    client, monkeypatch
):
    monkeypatch.setattr("app.main.run_task", lambda db, task: None)
    source_id = client.get("/api/sources").json()[0]["id"]

    response = client.post(
        "/api/tasks",
        json={
            "source_ids": [source_id],
            "start_date": "2026-08-01",
            "keyword_filter_enabled": True,
            "keyword_config": {
                "technical": [{"field": "集成电路", "keywords": "芯片"}],
                "industry_noun": [],
                "industry_verb": [{"keywords": "建设、规划"}],
            },
        },
    )

    assert response.status_code == 400
    assert "三维关键词配置不完整" in response.json()["detail"]


def test_queued_task_can_be_terminated(client, monkeypatch):
    monkeypatch.setattr("app.main.run_task", lambda db, task: None)
    source_id = client.get("/api/sources").json()[0]["id"]
    created = client.post(
        "/api/tasks", json={"source_ids": [source_id], "start_date": "2026-08-01"}
    ).json()

    response = client.post(f"/api/tasks/{created['id']}/terminate")

    assert response.status_code == 200
    assert response.json()["status"] == "terminated"
    assert response.json()["progress"] == 100
    assert response.json()["completed_at"] is not None
    logs = client.get(f"/api/tasks/{created['id']}/logs").json()
    assert "开始执行前已终止" in logs[-1]["message"]


def test_running_task_can_be_terminated(client, monkeypatch):
    def leave_running(db, task):
        task.status = "running"
        task.started_at = utc_now()
        db.commit()

    from app.models import utc_now

    monkeypatch.setattr("app.main.run_task", leave_running)
    source_id = client.get("/api/sources").json()[0]["id"]
    created = client.post(
        "/api/tasks", json={"source_ids": [source_id], "start_date": "2026-08-01"}
    ).json()

    response = client.post(f"/api/tasks/{created['id']}/terminate")

    assert response.status_code == 200
    assert response.json()["status"] == "terminated"
    assert response.json()["completed_at"] is not None
    repeated = client.post(f"/api/tasks/{created['id']}/terminate")
    assert repeated.status_code == 200
    assert repeated.json()["status"] == "terminated"


def test_finished_task_cannot_be_terminated(client, monkeypatch):
    def complete(db, task):
        task.status = "completed"
        task.completed_at = utc_now()
        db.commit()

    from app.models import utc_now

    monkeypatch.setattr("app.main.run_task", complete)
    source_id = client.get("/api/sources").json()[0]["id"]
    created = client.post(
        "/api/tasks", json={"source_ids": [source_id], "start_date": "2026-08-01"}
    ).json()

    response = client.post(f"/api/tasks/{created['id']}/terminate")

    assert response.status_code == 409
    assert "已经结束" in response.json()["detail"]


def test_empty_csv_and_xlsx_have_formal_columns(client):
    expected_headers = [label for _, label in EXPORT_COLUMNS]
    csv_response = client.get("/api/exports?format=csv")
    assert csv_response.status_code == 200
    csv_text = csv_response.content.decode("utf-8-sig")
    assert csv_text.splitlines()[0].split(",") == expected_headers
    assert "资讯类型" in csv_text

    xlsx_response = client.get("/api/exports?format=xlsx")
    assert xlsx_response.status_code == 200
    workbook = load_workbook(BytesIO(xlsx_response.content))
    sheet = workbook["结构化结果"]
    assert [cell.value for cell in sheet[1]] == expected_headers
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:J1"


def test_records_support_info_type_filter(client):
    response = client.get("/api/records?info_type=项目立项")
    assert response.status_code == 200
    assert response.json() == {"items": [], "total": 0}


def test_records_support_multiple_info_type_filters(client):
    from app.database import SessionLocal

    with SessionLocal() as db:
        db.add_all(
            [
                StructuredRecord(
                    region="华东",
                    company_name="甲公司",
                    info_type="项目立项",
                    source_name="测试来源",
                ),
                StructuredRecord(
                    region="华南",
                    company_name="乙公司",
                    info_type="项目签约",
                    source_name="测试来源",
                ),
                StructuredRecord(
                    region="华北",
                    company_name="丙公司",
                    info_type="项目竣工",
                    source_name="测试来源",
                ),
            ]
        )
        db.commit()

    response = client.get("/api/records?info_type=项目立项&info_type=项目签约")
    assert response.status_code == 200
    assert response.json()["total"] == 2
    assert {item["info_type"] for item in response.json()["items"]} == {
        "项目立项",
        "项目签约",
    }


def test_analytics_overview_builds_keywords_and_relations(client):
    from app.database import SessionLocal

    with SessionLocal() as db:
        db.add_all(
            [
                StructuredRecord(
                    region="合肥",
                    organization="高新区",
                    company_name="长鑫存储",
                    info_type="产能扩建",
                    project_name="晶圆产线",
                    details="DRAM 晶圆产能扩建",
                    source_name="测试来源",
                ),
                StructuredRecord(
                    region="合肥",
                    company_name="长鑫存储",
                    info_type="研发进展",
                    project_name="DRAM 芯片",
                    details="DRAM 芯片研发取得进展",
                    source_name="测试来源",
                ),
            ]
        )
        db.commit()

    client.put(
        "/api/settings/model",
        json={
            "base_url": "https://api.example.com",
            "model_name": "test-model",
            "keyword_config": [
                {"industry": "存储", "field": "先进封装", "keywords": "DRAM、晶圆"}
            ],
        },
    )
    response = client.get("/api/analytics/overview?region=合肥")
    assert response.status_code == 200
    data = response.json()
    assert data["summary"]["record_count"] == 2
    assert data["summary"]["entity_count"] >= 4
    assert any(
        node["name"] == "长鑫存储" and node["value"] == 2
        for node in data["graph"]["nodes"]
    )
    assert any(edge["value"] >= 1 for edge in data["graph"]["edges"])
    assert any(item["name"] == "产能扩建" for item in data["info_types"])
    assert {item["text"] for item in data["keywords"]} == {"DRAM", "晶圆"}


def test_analytics_merges_same_entity_across_record_fields(client):
    from app.database import SessionLocal

    with SessionLocal() as db:
        db.add(
            StructuredRecord(
                region="海外",
                organization="Cerebras Systems",
                company_name="Cerebras Systems",
                info_type="经营动态",
                source_name="测试来源",
            )
        )
        db.commit()

    response = client.get("/api/analytics/overview")
    assert response.status_code == 200
    nodes = [
        node
        for node in response.json()["graph"]["nodes"]
        if node["name"] == "Cerebras Systems"
    ]
    assert len(nodes) == 1
    assert nodes[0]["category"] == "企业"


def test_history_full_text_search_and_manual_structure_guard(client):
    from app.database import SessionLocal
    from app.models import Source

    with SessionLocal() as db:
        source = db.query(Source).first()
        article = RawArticle(
            source_id=source.id,
            canonical_url="https://example.com/chip-project",
            title="晶圆厂扩建计划",
            published_text="2026年8月20日",
            body="华东晶圆制造项目新增产线",
            content_hash="a" * 64,
            status="pending",
        )
        db.add(article)
        db.flush()
        db.add(
            StructuredRecord(
                article_id=article.id,
                region="华东",
                company_name="晶圆公司",
                info_type="产能扩建",
                investment_amount="未披露",
                project_name="新增产线",
                source_name=source.name,
                original_url=article.canonical_url,
                details="建设先进晶圆产线",
            )
        )
        article_id = article.id
        db.commit()

    raw = client.get("/api/articles?q=新增产线")
    assert raw.status_code == 200
    assert raw.json()["total"] == 1
    assert raw.json()["items"][0]["record_count"] == 1
    assert raw.json()["items"][0]["collected_at"].endswith("+08:00")

    detail = client.get(f"/api/articles/{article_id}")
    assert detail.status_code == 200
    assert detail.json()["collected_at"].endswith("+08:00")

    records = client.get("/api/records?q=先进晶圆")
    assert records.status_code == 200
    assert records.json()["total"] == 1

    response = client.post(f"/api/articles/{article_id}/structure")
    assert response.status_code == 409
    assert "已经完成结构化" in response.json()["detail"]


def test_manual_structure_requires_configured_model(client):
    from app.database import SessionLocal
    from app.models import Source

    with SessionLocal() as db:
        source = db.query(Source).first()
        article = RawArticle(
            source_id=source.id,
            canonical_url="https://example.com/unstructured",
            title="待处理原文",
            body="正文",
            content_hash="b" * 64,
            status="pending",
        )
        db.add(article)
        db.commit()
        article_id = article.id

    response = client.post(f"/api/articles/{article_id}/structure")
    assert response.status_code == 409
    assert "API Key" in response.json()["detail"]


def test_manual_structure_ignores_auto_structure_switch(client, monkeypatch):
    from app.database import SessionLocal
    from app.models import Source

    setting = client.put(
        "/api/settings/model",
        json={
            "base_url": "https://api.example.com",
            "model_name": "test-model",
            "api_key": "sk-secret1234",
            "enabled": False,
        },
    )
    assert setting.status_code == 200

    with SessionLocal() as db:
        source = db.query(Source).first()
        article = RawArticle(
            source_id=source.id,
            canonical_url="https://example.com/manual",
            title="手动结构化原文",
            body="正文",
            content_hash="c" * 64,
            status="pending",
        )
        db.add(article)
        db.commit()
        article_id = article.id

    def fake_structure(db, article, setting):
        article.status = "completed"
        article.model_name = setting.model_name
        return 1

    monkeypatch.setattr("app.main.structure_article", fake_structure)
    response = client.post(f"/api/articles/{article_id}/structure")
    assert response.status_code == 200
    assert response.json() == {
        "article_id": article_id,
        "created_count": 1,
        "status": "completed",
    }


def test_model_setting_masks_secret(client):
    response = client.put(
        "/api/settings/model",
        json={
            "base_url": "https://api.example.com",
            "model_name": "test-model",
            "api_key": "sk-secret1234",
            "baidu_api_key": "baidu-secret5678",
            "anysearch_api_key": "as_sk-secret5678",
            "request_headers": [
                {"key": "X-API-Version", "value": "2026-08-01"},
            ],
            "enabled": True,
        },
    )
    assert response.status_code == 200
    assert response.json()["has_api_key"] is True
    assert response.json()["api_key_hint"].endswith("1234")
    assert response.json()["has_baidu_api_key"] is True
    assert response.json()["baidu_api_key_hint"].endswith("5678")
    assert response.json()["has_anysearch_api_key"] is True
    assert response.json()["anysearch_api_key_hint"].endswith("5678")
    assert response.json()["request_headers"] == [
        {"key": "X-API-Version", "value": "2026-08-01"},
    ]
    assert "secret" not in response.text
    assert "api_key" not in client.get("/api/settings/model").json()


def test_model_setting_rejects_duplicate_request_headers(client):
    response = client.put(
        "/api/settings/model",
        json={
            "base_url": "https://api.example.com",
            "model_name": "test-model",
            "request_headers": [
                {"key": "X-Tenant", "value": "first"},
                {"key": "x-tenant", "value": "second"},
            ],
        },
    )
    assert response.status_code == 422


def test_background_task_rolls_back_before_recording_failure(client, monkeypatch):
    from app.database import SessionLocal
    from app.main import _run_task_background
    from app.models import CollectionTask, Source, TaskLog

    with SessionLocal() as db:
        source = db.query(Source).first()
        task = CollectionTask(
            status="queued",
            start_date=date(2026, 8, 1),
            source_ids_json=f"[{source.id}]",
            source_snapshot_json="[]",
        )
        db.add(task)
        db.commit()
        task_id = task.id

    def fail_after_flush(db, task):
        db.add(
            Source(name="冲突来源", base_url="https://example.com", config_json="{}")
        )
        db.flush()
        db.add(
            Source(name="冲突来源", base_url="https://example.com", config_json="{}")
        )
        db.flush()

    monkeypatch.setattr("app.main.run_task", fail_after_flush)
    _run_task_background(task_id)

    with SessionLocal() as db:
        task = db.get(CollectionTask, task_id)
        logs = db.query(TaskLog).filter(TaskLog.task_id == task_id).all()
        assert task.status == "failed"
        assert task.completed_at is not None
        assert any("任务失败" in log.message for log in logs)


def test_source_rejects_cross_domain_entry(client):
    response = client.post(
        "/api/sources",
        json={
            "name": "越界来源",
            "base_url": "https://example.com",
            "config": {
                "entry_urls": ["https://other.example/news"],
                "article_url_pattern": "/news/",
                "selectors": {"title": "h1", "date": ".date", "content": ".content"},
            },
        },
    )
    assert response.status_code == 422


def test_audit_export_has_trace_columns(client):
    response = client.get("/api/exports?format=xlsx&columns=audit")
    workbook = load_workbook(BytesIO(response.content))
    headers = [cell.value for cell in workbook["结构化结果"][1]]
    assert "字段证据" in headers
    assert "任务ID" in headers
